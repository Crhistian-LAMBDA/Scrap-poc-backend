from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from typing import Any

from django.http import HttpResponse
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status

from .services.seguros_bolivar_session import BolivarResult, SegurosBolivarSession


def _normalize_cookie_header(raw: str) -> str:
	value = (raw or "").strip()
	if not value:
		return ""

	# Soportar: "Cookie: a=b; c=d" o solo "a=b; c=d"
	if value.lower().startswith("cookie:"):
		value = value.split(":", 1)[1].strip()

	# Aceptar cookies pegadas en varias líneas (una por cookie)
	value = value.replace("\r", "")
	value = "; ".join([p.strip() for p in value.replace("\n", ";").split(";") if p.strip()])

	return value


def _normalize_radicados(value: Any) -> list[str]:
	if value is None:
		return []

	if isinstance(value, str):
		# Soportar textarea con separadores comunes
		candidates = (
			value.replace("\r", "\n")
			.replace(",", "\n")
			.replace(";", "\n")
			.split("\n")
		)
		radicados = [c.strip() for c in candidates if c.strip()]
	elif isinstance(value, list):
		radicados = [str(v).strip() for v in value if str(v).strip()]
	else:
		radicados = [str(value).strip()] if str(value).strip() else []

	# Deduplicar manteniendo orden
	seen: set[str] = set()
	unique: list[str] = []
	for r in radicados:
		if r not in seen:
			unique.append(r)
			seen.add(r)

	return unique


def _to_result_dict(result: BolivarResult) -> dict[str, Any]:
	return {
		"radicado": result.radicado,
		"ok": result.ok,
		"estado_raw": result.estado_raw,
		"estado_normalizado": result.estado_normalizado,
		"asegurado": result.asegurado,
		"consulted_at": result.consulted_at,
		"error": result.error,
	}


def _export_xlsx_clean(rows: list[dict[str, Any]]) -> bytes:
	from openpyxl import Workbook
	from openpyxl.utils import get_column_letter

	headers = [
		"Radicado",
		"Estado (portal)",
		"Estado (normalizado)",
		"Asegurado",
		"Fecha de consulta",
	]

	wb = Workbook()
	ws = wb.active
	ws.title = "radicados"

	ws.append(headers)
	for r in rows:
		ws.append(
			[
				r.get("radicado") or "",
				r.get("estado_raw") or "",
				r.get("estado_normalizado") or "",
				r.get("asegurado") or "",
				r.get("consulted_at") or "",
			]
		)

	# Auto-ajuste simple de columnas (por longitud de texto)
	for col_idx in range(1, len(headers) + 1):
		max_len = 0
		for row in ws.iter_rows(min_row=1, max_col=col_idx, min_col=col_idx):
			cell = row[0]
			val = "" if cell.value is None else str(cell.value)
			max_len = max(max_len, len(val))
		ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

	buffer = BytesIO()
	wb.save(buffer)
	return buffer.getvalue()


def _export_pdf_clean(rows: list[dict[str, Any]], generated_at: str) -> bytes:
	from reportlab.lib import colors
	from reportlab.lib.pagesizes import letter
	from reportlab.lib.styles import getSampleStyleSheet
	from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

	buffer = BytesIO()
	doc = SimpleDocTemplate(buffer, pagesize=letter, title="Resultado consulta de radicados")
	styles = getSampleStyleSheet()

	story = []
	story.append(Paragraph("Resultado consulta de radicados", styles["Title"]))
	story.append(Spacer(1, 6))
	story.append(Paragraph(f"Fecha de generación: {generated_at}", styles["Normal"]))
	story.append(Spacer(1, 12))

	data = [
		[
			"Radicado",
			"Estado (portal)",
			"Estado (normalizado)",
			"Asegurado",
			"Fecha consulta",
		]
	]

	for r in rows:
		data.append(
			[
				r.get("radicado") or "",
				r.get("estado_raw") or "",
				r.get("estado_normalizado") or "",
				r.get("asegurado") or "",
				r.get("consulted_at") or "",
			]
		)

	table = Table(data, repeatRows=1)
	table.setStyle(
		TableStyle(
			[
				("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
				("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
				("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
				("FONTSIZE", (0, 0), (-1, -1), 9),
				("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
				("VALIGN", (0, 0), (-1, -1), "TOP"),
				("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
			]
		)
	)

	story.append(table)
	doc.build(story)

	return buffer.getvalue()


def _sanitize_export_rows(results: Any) -> list[dict[str, Any]]:
	"""Construye filas exportables SOLO con campos funcionales (sin debug/trazas)."""
	if not isinstance(results, list):
		return []

	rows: list[dict[str, Any]] = []
	for item in results:
		if not isinstance(item, dict):
			continue
		rows.append(
			{
				"radicado": (item.get("radicado") or "").strip(),
				"estado_raw": (item.get("estado_raw") or "").strip(),
				"estado_normalizado": (item.get("estado_normalizado") or "").strip(),
				"asegurado": (item.get("asegurado") or "").strip(),
				"consulted_at": (item.get("consulted_at") or "").strip(),
			}
		)

	# Filtrar filas vacías
	rows = [r for r in rows if r.get("radicado")]
	return rows


@api_view(["POST"])
def bolivar_radicados(request):
	"""Endpoint único para consulta por radicados usando cookie de sesión pegada.

	Request JSON:
	{
	  "cookie": "JSESSIONID=...; ..."   // o "Cookie: ..."
	  "radicados": ["123", "456"]     // o string multilinea
	  "export": null|"xlsx"|"pdf"
	}
	"""

	cookie_header = _normalize_cookie_header(request.data.get("cookie") or "")
	radicados = _normalize_radicados(request.data.get("radicados"))
	use_server_auth = bool(request.data.get("use_server_auth"))

	if not cookie_header and not use_server_auth:
		return Response(
			{
				"detail": "Falta 'cookie' o 'use_server_auth=true'. Este sistema solo consulta; no ejecuta acciones críticas.",
			},
			status=status.HTTP_400_BAD_REQUEST,
		)

	if not radicados:
		return Response(
			{"detail": "Falta 'radicados' (lista o texto con radicados)."},
			status=status.HTTP_400_BAD_REQUEST,
		)

	fetched_at = datetime.now(timezone.utc).isoformat()
	results: list[BolivarResult] = []

	# Inicializa sesión (inyección de cookie) y consulta cada radicado.
	try:
		session = SegurosBolivarSession(
			cookie_header=cookie_header or None,
			use_server_auth=use_server_auth,
		)
	except Exception as exc:  # noqa: BLE001
		return Response(
			{"detail": f"No se pudo inicializar sesión: {exc}"},
			status=status.HTTP_400_BAD_REQUEST,
		)

	for radicado in radicados:
		consulted_at = datetime.now(timezone.utc).isoformat()
		try:
			estado_raw, estado_normalizado, asegurado = session.get_info_for_radicado(
				radicado
			)
			results.append(
				BolivarResult(
					radicado=radicado,
					ok=True,
					estado_raw=estado_raw,
					estado_normalizado=estado_normalizado,
					asegurado=asegurado,
					consulted_at=consulted_at,
					error=None,
				)
			)
		except NotImplementedError as exc:
			results.append(
				BolivarResult(
					radicado=radicado,
					ok=False,
					estado_raw=None,
					estado_normalizado="NO ENCONTRADO",
					asegurado=None,
					consulted_at=consulted_at,
					error=str(exc),
				)
			)
		except Exception as exc:  # noqa: BLE001
			results.append(
				BolivarResult(
					radicado=radicado,
					ok=False,
					estado_raw=None,
					estado_normalizado="NO ENCONTRADO",
					asegurado=None,
					consulted_at=consulted_at,
					error=str(exc),
				)
			)

	return Response(
		{
			"fetched_at": fetched_at,
			"count": len(results),
			"results": [_to_result_dict(r) for r in results],
		}
	)


@api_view(["POST"])
def bolivar_radicados_export(request):
	"""Genera XLSX/PDF a partir de resultados YA consultados (sin re-scrapear).

	Request JSON:
	{
	  "format": "xlsx" | "pdf",
	  "results": [ {radicado, estado_raw, estado_normalizado, asegurado?, consulted_at?}, ... ]
	}
	"""

	export_format = (request.data.get("format") or "").strip().lower()
	rows = _sanitize_export_rows(request.data.get("results"))

	if export_format not in {"xlsx", "pdf"}:
		return Response(
			{"detail": "Formato inválido. Usa 'xlsx' o 'pdf'."},
			status=status.HTTP_400_BAD_REQUEST,
		)
	if not rows:
		return Response(
			{"detail": "No hay resultados para exportar. Primero consulta radicados."},
			status=status.HTTP_400_BAD_REQUEST,
		)

	timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
	generated_at = datetime.now(timezone.utc).isoformat()

	if export_format == "xlsx":
		payload = _export_xlsx_clean(rows)
		content_type = (
			"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
		)
		filename = f"radicados_{timestamp}.xlsx"
	else:
		payload = _export_pdf_clean(rows, generated_at=generated_at)
		content_type = "application/pdf"
		filename = f"radicados_{timestamp}.pdf"

	response = HttpResponse(payload, content_type=content_type)
	response["Content-Disposition"] = f'attachment; filename="{filename}"'
	return response
